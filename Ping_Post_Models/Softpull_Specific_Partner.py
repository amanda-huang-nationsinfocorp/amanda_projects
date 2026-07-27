"""
m0_conversion_by_attribute_compare.py
-------------------------------------
Runs the Bloodhound query for the WHOLE VERTICAL (no partner filter) and writes
m0_conversion_by_attribute_compare.xlsx.

Same layout/analysis as before, plus two comparison columns so you can read the
vertical vs. one partner side-by-side within the SAME buckets:

    m0_948018                = partner 948018's m0 count in that bucket
    conversion_rate_948018   = partner 948018's m0 rate  in that bucket
                               (blank where 948018 has no rows in the bucket)

Bins are defined once on the whole-vertical data, then partner 948018's rows are
scored into those exact buckets — so every row compares like-for-like.

SETUP:
    pip install "snowflake-connector-python[pandas]" openpyxl pandas numpy
    export SNOWFLAKE_ACCOUNT=... SNOWFLAKE_USER=... SNOWFLAKE_PASSWORD=... \
           SNOWFLAKE_WAREHOUSE=... SNOWFLAKE_ROLE=... SNOWFLAKE_DATABASE=ANALYTICS
RUN:
    python m0_conversion_by_attribute_compare.py
"""

#%%
from __future__ import annotations
import os, sys, warnings
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule

warnings.filterwarnings("ignore")

# ======================================================================
# SCOPE
# ======================================================================
PARTNER_ID         = None     # None = whole vertical (no partner_id filter)
COMPARE_PARTNER_ID = 948018   # partner broken out into the *_948018 columns

# ======================================================================
# CONFIG
# ======================================================================
TARGET_COL   = "label"
POSITIVE_VAL = "m0"
N_BINS       = 7
MAX_CARD     = 30
NUMERIC_COERCE_MIN = 0.80
MIN_BUCKET_N = 1
ARIAL = "Arial"
OUT_PATH = "m0_conversion_by_attribute_compare.xlsx"

M0_CMP_COL = f"m0_{COMPARE_PARTNER_ID}"
CR_CMP_COL = f"conversion_rate_{COMPARE_PARTNER_ID}"

EXCLUDE_COLS = {
    "jluvr", "order_id", "partner_id", TARGET_COL,
    "lead_user_agent", "employer", "bank_account_bank_name",
}

# ======================================================================
# QUERY  (your query verbatim through the `final` CTE; final SELECT appended)
# ======================================================================
BASE_QUERY = r"""
WITH pingpost AS (
    SELECT
        jluvr, partner_id, order_id, request_datetime, seller_name,
        CASE
            WHEN seller_name = 'ZP' THEN '233 Zero Parallel LLC'
            WHEN seller_name = 'PingYo' THEN '1011 PingYo'
            WHEN seller_name = 'Bume' THEN '504 Bume Intl'
            ELSE seller_name
        END AS mapped_seller_name,
        label, seller_minimum_price, lead_user_agent, loan_amount, loan_purpose,
        finance_current_unsecured_debt_amount, applicant_employment_type,
        applicant_income_net_monthly_amount, applicant_employment_length,
        applicant_income_last_pay_day, applicant_income_next_pay_day,
        applicant_date_of_birth, applicant_current_address_state,
        applicant_current_address_residence_status, applicant_current_address_years,
        bank_account_bank_name, bank_account_years, consent_marketing_email
    FROM ANALYTICS.BLOODHOUND_PROJECT.PINGPOST_DATA_PARSED
),
soft_pull AS (
    SELECT
        jluvr, to_date(inquiry_date) as inquiry_date, risk_score,
        YEAR(CURRENT_DATE()) - BIRTH_YEAR as age, employer,
        IFF('12' IN (negative_factor_1, negative_factor_2, negative_factor_3, negative_factor_4), 1, 0) AS has_negative_factor_12,
        IFF('63' IN (negative_factor_1, negative_factor_2, negative_factor_3, negative_factor_4), 1, 0) AS has_negative_factor_63,
        IFF('07' IN (negative_factor_1, negative_factor_2, negative_factor_3, negative_factor_4), 1, 0) AS has_negative_factor_07,
        IFF('95' IN (negative_factor_1, negative_factor_2, negative_factor_3, negative_factor_4), 1, 0) AS has_negative_factor_95,
        IFF('04' IN (negative_factor_1, negative_factor_2, negative_factor_3, negative_factor_4), 1, 0) AS has_negative_factor_04,
        IFF('P05' IN (positive_factor_1, positive_factor_2, positive_factor_3, positive_factor_4), 1, 0) AS has_positive_factor_P05,
        IFF('P34' IN (positive_factor_1, positive_factor_2, positive_factor_3, positive_factor_4), 1, 0) AS has_positive_factor_P34,
        IFF('P08' IN (positive_factor_1, positive_factor_2, positive_factor_3, positive_factor_4), 1, 0) AS has_positive_factor_P08,
        IFF('P04' IN (positive_factor_1, positive_factor_2, positive_factor_3, positive_factor_4), 1, 0) AS has_positive_factor_P04,
        IFF('P95' IN (positive_factor_1, positive_factor_2, positive_factor_3, positive_factor_4), 1, 0) AS has_positive_factor_P95
    FROM ANALYTICS.BLOODHOUND_PROJECT.SOFTPULL_SERVICE_PRODUCT_RESPONSE
),
trade_lines AS (
    SELECT
        jluvr, count(*) as total_account_count,
        count_if(account_current_status = 'Open') as total_open_account,
        COUNT_IF(DATEDIFF('month', account_open_date, CURRENT_DATE()) <= 6) AS account_opened_within_6_month,
        count_if(account_type in ('Revolving', 'Credit Line', 'Open Account', 'Mortgage')) as revolving_account_count,
        count_if(account_type = 'Installment') as installment_account_count,
        count_if(account_type = 'Colletion') as collection_account_count,
        sum(high_balance) as total_high_balance, sum(current_balance) as total_current_balance,
        avg(credit_limit) as credit_limit_avg,
        sum(case when account_current_status = 'Open' then credit_limit else null end) as open_credit_limit_total,
        div0(sum(monthly_payment), sum(case when account_current_status = 'Open' then credit_limit else null end)) as credit_utilization_ratio,
        max(DATEDIFF('month', account_open_date, CURRENT_DATE()) / 12.0) as earliest_account_open_year,
        min(DATEDIFF('month', account_open_date, CURRENT_DATE()) / 12.0) as latest_account_open_year,
        count_if(account_status = 'Derogatory') as derogatory_account_count,
        div0(count_if(account_status = 'Derogatory'), count(*)) as derogatory_account_ratio,
        div0(count_if(account_status = 'Derogatory' and account_current_status = 'Closed'),count_if(account_status = 'Derogatory')) as closed_derogatory_account_ratio,
        count_if(account_status = 'Paid') as paid_account_count,
        count_if(account_status in ('Refinanced', 'Transferred')) as transferred_account_count,
        div0(count_if(payment_status = 'Current'),count(*)) as on_time_payment_current_ratio,
        count_if(open_account_type = 'Unsecured loan') as unsecured_loan_account_count,
        count_if(open_account_type = 'Secured loan') as secured_loan_account_count,
        count_if(open_account_type = 'Educational') as educational_account_count,
        count_if(lower(open_account_type) like 'credit card') as credit_card_account_count,
        count_if(open_account_type = 'Charge account') as charge_account_count,
        count_if(creditor_category = 'BNPL_Fintech_Advance') as fintech_creditor_count,
        count_if(open_account_type = 'Unsecured loan' and creditor_category = 'BNPL_Fintech_Advance') as unsecured_loan_fintech_count,
        count_if(open_account_type = 'Unsecured loan' and creditor_category = 'Personal_Installment_Payday_Loan') as unsecured_loan_ploan_count,
        sum(case when account_current_status = 'Open' then monthly_payment end) as current_monthly_payment_total,
        div0(sum(case when account_current_status = 'Open' then monthly_payment end), sum(current_balance)) as payment_burden_ratio,
        avg(monthly_payment) as historic_monthly_payment_avg,
        avg(month_reviewed) as month_reviewed_avg,
        sum(late_30_count) as late_30_total, sum(late_60_count) as late_60_total,
        sum(late_90_count) as late_90_total, sum(amount_past_due) as past_due_amount_total,
        div0(sum(amount_past_due), sum(current_balance)) as past_due_to_current_balance_ratio
    FROM ANALYTICS.BLOODHOUND_PROJECT.SOFTPULL_TRADE_LINE
    GROUP BY 1
),
inquiry AS (
    SELECT
        jluvr, count(*) as inquiry_total,
        COUNT_IF(DATEDIFF('month', inquiry_date, CURRENT_DATE()) <= 6) AS inquiry_within_6_month,
        div0(count_if(industry_code = 'Bank'),count(*)) as bank_inquiry_ratio,
        div0(count_if(industry_code = 'Finance/Personal'),count(*)) as personal_inquiry_ratio
    FROM ANALYTICS.BLOODHOUND_PROJECT.SOFTPULL_INQUIRY
    GROUP BY 1
),
public_record AS (
    SELECT
        jluvr, count(*) as bankcrupcy_total,
        min(DATEDIFF('month', public_record_date, CURRENT_DATE()) / 12.0) as most_recent_record_year,
        div0(count_if(status = 'Discharged'), count(*)) as bankcrupty_discharged_ratio
    FROM ANALYTICS.BLOODHOUND_PROJECT.SOFTPULL_PUBLIC_RECORD
    WHERE classification = 'Bankruptcy'
    GROUP BY 1
),
daily_partner_stats AS (
    SELECT
        super_partner_id_name, TO_DATE(order_date) AS order_date,
        COUNT(*) AS daily_total_count, COUNT_IF(is_m0 = 1) AS m0_count,
        COUNT_IF(bin_tier = 'Tier1') AS bin_tier_1_count,
        COUNT_IF(bin_tier = 'Tier2') AS bin_tier_2_count,
        COUNT_IF(bin_tier = 'Tier3') AS bin_tier_3_count,
        COUNT_IF(bin_tier = 'Tier4') AS bin_tier_4_count,
        COUNT_IF(bin_tier = 'Blacklist') AS bin_tier_Blacklist_count
    FROM DBT_PROD.ANALYTICS.FCT_TRANSACTION_INVOICE_ORDER_ITEM AS orders
    LEFT JOIN DBT_PROD.ANALYTICS.stg_dbt_prod_analytics_bin_tier_matching AS bin_tiers
        ON orders.vertical_type_bin = bin_tiers.vertical_type_bin
    WHERE super_partner_id IN (504, 1011, 233)
    GROUP BY 1, 2
),
partner_past_3d_stats AS (
    SELECT
        p.jluvr, AVG(d.m0_count) AS m0_past_3_days,
        DIV0(SUM(d.bin_tier_1_count), SUM(d.daily_total_count)) AS bin_tier_1_perc_past_3_days,
        DIV0(SUM(d.bin_tier_2_count), SUM(d.daily_total_count)) AS bin_tier_2_perc_past_3_days,
        DIV0(SUM(d.bin_tier_3_count), SUM(d.daily_total_count)) AS bin_tier_3_perc_past_3_days,
        DIV0(SUM(d.bin_tier_4_count), SUM(d.daily_total_count)) AS bin_tier_4_perc_past_3_days,
        DIV0(SUM(d.bin_tier_Blacklist_count), SUM(d.daily_total_count)) AS bin_tier_Blacklist_perc_past_3_days
    FROM pingpost p
    LEFT JOIN daily_partner_stats d
        ON p.mapped_seller_name = d.super_partner_id_name
        AND d.order_date >= TO_DATE(p.request_datetime) - 3
        AND d.order_date <  TO_DATE(p.request_datetime)
    GROUP BY p.jluvr
),
final AS (
    SELECT
        p.jluvr, p.partner_id, p.order_id, p.request_datetime, p.seller_name, p.label,
        p.seller_minimum_price, p.lead_user_agent, p.loan_amount, p.loan_purpose,
        p.finance_current_unsecured_debt_amount, p.applicant_employment_type,
        p.applicant_income_net_monthly_amount, p.applicant_employment_length,
        p.applicant_income_last_pay_day, p.applicant_income_next_pay_day,
        p.applicant_date_of_birth, p.applicant_current_address_state,
        p.applicant_current_address_residence_status, p.applicant_current_address_years,
        p.bank_account_bank_name, p.bank_account_years, p.consent_marketing_email,
        sp.inquiry_date, sp.risk_score, sp.age, sp.employer,
        tl.total_account_count, tl.total_open_account, tl.revolving_account_count,
        tl.installment_account_count, tl.collection_account_count, tl.total_high_balance,
        tl.total_current_balance, tl.credit_utilization_ratio, tl.earliest_account_open_year,
        tl.latest_account_open_year, tl.derogatory_account_count, tl.derogatory_account_ratio,
        tl.closed_derogatory_account_ratio, tl.paid_account_count, tl.transferred_account_count,
        tl.on_time_payment_current_ratio, tl.unsecured_loan_account_count,
        tl.secured_loan_account_count, tl.educational_account_count, tl.credit_card_account_count,
        tl.charge_account_count, tl.current_monthly_payment_total, tl.payment_burden_ratio,
        tl.historic_monthly_payment_avg, tl.month_reviewed_avg, tl.late_30_total,
        tl.late_60_total, tl.late_90_total, tl.past_due_amount_total,
        tl.past_due_to_current_balance_ratio, iq.inquiry_total, iq.inquiry_within_6_month,
        iq.bank_inquiry_ratio, iq.personal_inquiry_ratio, pr.bankcrupcy_total,
        pr.most_recent_record_year, pr.bankcrupty_discharged_ratio, p3d.m0_past_3_days,
        p3d.bin_tier_1_perc_past_3_days, p3d.bin_tier_2_perc_past_3_days,
        p3d.bin_tier_3_perc_past_3_days, p3d.bin_tier_4_perc_past_3_days,
        p3d.bin_tier_Blacklist_perc_past_3_days
    FROM pingpost p
    LEFT JOIN soft_pull sp USING(jluvr)
    LEFT JOIN trade_lines tl USING(jluvr)
    LEFT JOIN inquiry iq USING(jluvr)
    LEFT JOIN public_record pr USING(jluvr)
    LEFT JOIN partner_past_3d_stats p3d USING(jluvr)
    WHERE sp.risk_score IS NOT NULL
)
"""

def build_query(partner_id):
    where = f"\nWHERE partner_id = {int(partner_id)}" if partner_id is not None else ""
    return BASE_QUERY + "\nSELECT * FROM final" + where


# ======================================================================
# SNOWFLAKE FETCH
# ======================================================================
def fetch(partner_id) -> pd.DataFrame:
    import snowflake.connector
    conn = snowflake.connector.connect(
    user='BITEAM',
    password='B1sense@22',
    account='YXBYZCG-MVA06208',
    database="DBT_PROD.PUBLIC",
    )
    try:
        cur = conn.cursor()
        cur.execute(build_query(partner_id))
        df = cur.fetch_pandas_all()
    finally:
        conn.close()
    df.columns = [c.lower() for c in df.columns]
    return df


# ======================================================================
# ANALYSIS
# ======================================================================
def _coerce(series: pd.Series):
    nn = series.notna().sum()
    if nn == 0:
        return series, "categorical"
    num = pd.to_numeric(series, errors="coerce")
    if num.notna().sum() >= NUMERIC_COERCE_MIN * nn:
        return num, "numeric"
    dt = pd.to_datetime(series, errors="coerce")
    if dt.notna().sum() >= NUMERIC_COERCE_MIN * nn:
        return dt, "datetime"
    return series, "categorical"


def _make_buckets(series: pd.Series):
    conv, kind = _coerce(series)
    n_unique = conv.nunique(dropna=True)
    if kind in ("numeric", "datetime") and n_unique > 1:
        vals = conv.astype("int64") if kind == "datetime" else conv
        try:
            bucket = pd.qcut(vals, q=N_BINS, duplicates="drop")
        except (ValueError, IndexError):
            try:
                bucket = pd.cut(vals, bins=min(N_BINS, n_unique))
            except (ValueError, IndexError):
                bucket = conv.astype("object")
        bucket = bucket.astype("object")
        if kind == "datetime":
            def relabel(iv):
                if pd.isna(iv) or not hasattr(iv, "left"):
                    return iv
                return f"{pd.Timestamp(int(iv.left)).date()} - {pd.Timestamp(int(iv.right)).date()}"
            bucket = bucket.map(relabel)
        return bucket.where(conv.notna(), "(Missing)"), kind

    bucket = conv.astype("object").where(conv.notna(), "(Missing)")
    if n_unique > MAX_CARD:
        top = conv.value_counts().head(MAX_CARD - 1).index
        bucket = bucket.where(bucket.isin(top) | (bucket == "(Missing)"), "(Other)")
    return bucket, kind


def _sort_key(label):
    if label in ("(Missing)", "(Other)"):
        return (2, str(label))
    if isinstance(label, pd.Interval):
        return (0, label.left)
    return (1, str(label))


def analyze(df: pd.DataFrame):
    df.columns = [c.lower() for c in df.columns]
    if TARGET_COL not in df.columns:
        sys.exit(f"Target column '{TARGET_COL}' not found. Columns: {list(df.columns)}")

    tgt = df[TARGET_COL].astype("string").str.strip().str.lower()
    is_pos = (tgt == POSITIVE_VAL.lower()).fillna(False).astype(int)
    overall = float(is_pos.mean())

    # partner-948018 mask for the comparison columns
    if "partner_id" in df.columns:
        pmask = (pd.to_numeric(df["partner_id"], errors="coerce") == COMPARE_PARTNER_ID).astype(int)
    else:
        print(f"[warn] no 'partner_id' column — {M0_CMP_COL}/{CR_CMP_COL} will be blank.")
        pmask = pd.Series(0, index=df.index)

    p_rows = int(pmask.sum())
    p_m0   = int((is_pos * pmask).sum())
    p_overall = (p_m0 / p_rows) if p_rows else np.nan

    rows = []
    for col in [c for c in df.columns if c not in EXCLUDE_COLS]:
        if df[col].nunique(dropna=True) == 0:
            continue
        buckets, kind = _make_buckets(df[col])
        g = pd.DataFrame({"bucket": buckets,
                          "is_pos": is_pos.values,
                          "p": pmask.values})
        g["pos_p"] = g["is_pos"] * g["p"]
        agg = (g.groupby("bucket", dropna=False)
                 .agg(n=("is_pos", "size"), m0=("is_pos", "sum"),
                      n_p=("p", "sum"), m0_p=("pos_p", "sum")).reset_index())
        agg = agg[agg["n"] >= MIN_BUCKET_N]
        agg["attribute"]       = col
        agg["type"]            = kind
        agg["share"]           = agg["n"] / len(df)
        agg["m0"]              = agg["m0"].astype(int)
        agg["conversion_rate"] = agg["m0"] / agg["n"]
        agg[M0_CMP_COL]        = agg["m0_p"].astype(int)
        agg[CR_CMP_COL]        = np.where(agg["n_p"] > 0, agg["m0_p"] / agg["n_p"], np.nan)
        agg["lift_vs_overall"] = agg["conversion_rate"] / overall if overall else np.nan
        agg = agg.sort_values("bucket", key=lambda s: s.map(_sort_key))
        rows.append(agg)

    out = pd.concat(rows, ignore_index=True)
    out["bucket"] = out["bucket"].astype(str)
    out = out[["attribute", "type", "bucket", "n", "share", "m0", M0_CMP_COL,
               "conversion_rate", CR_CMP_COL, "lift_vs_overall"]]
    out = out.round({"share": 4, "conversion_rate": 4, CR_CMP_COL: 4, "lift_vs_overall": 4})

    meta = dict(overall=overall, n_rows=len(df), n_m0=int(is_pos.sum()),
                p_rows=p_rows, p_m0=p_m0, p_overall=p_overall)
    return out, meta


# ======================================================================
# WORKBOOK
# ======================================================================
def build_xlsx(df_res, meta, scope, source, out_path):
    navy = PatternFill("solid", fgColor="1F3864")
    thin = Side(style="thin", color="D9D9D9")
    grp  = Side(style="medium", color="808080")
    cid  = COMPARE_PARTNER_ID

    wb = Workbook()
    ws = wb.active; ws.title = "Conversion by attribute"
    ws["A1"] = f"m0 Conversion by Attribute — {scope} (vs Partner {cid})"
    ws["A1"].font = Font(name=ARIAL, size=14, bold=True)
    po = f"{meta['p_overall']:.4f}" if meta["p_rows"] else "n/a"
    ws["A2"] = (f"Vertical m0 rate: {meta['overall']:.4f} (rows {meta['n_rows']:,}, m0 {meta['n_m0']:,})   |   "
                f"Partner {cid} m0 rate: {po} (rows {meta['p_rows']:,}, m0 {meta['p_m0']:,})   |   "
                f"Source: {source}.")
    ws["A2"].font = Font(name=ARIAL, size=9, italic=True, color="555555")
    ws["A3"] = (f"share & all conversion_rate columns are raw proportions (0–1).  lift = vertical conversion_rate ÷ vertical overall.  "
                f"{M0_CMP_COL} / {CR_CMP_COL} = partner {cid}'s count & rate within the SAME bucket (blank = no {cid} rows in that bucket).")
    ws["A3"].font = Font(name=ARIAL, size=9, italic=True, color="555555")

    headers = ["attribute","type","bucket","n","share","m0", M0_CMP_COL,
               "conversion_rate", CR_CMP_COL, "lift_vs_overall"]
    HDR = 5
    for j,h in enumerate(headers, 1):
        c = ws.cell(HDR, j, h)
        c.font = Font(name=ARIAL, bold=True, color="FFFFFF"); c.fill = navy
        c.alignment = Alignment(horizontal="center", vertical="center")

    numfmt = {4:"#,##0", 5:"0.0000", 6:"#,##0", 7:"#,##0", 8:"0.0000", 9:"0.0000", 10:"0.00"}
    prev, r = None, HDR + 1
    for _, row in df_res.iterrows():
        new = row["attribute"] != prev
        for j,h in enumerate(headers, 1):
            v = row[h]
            if h == CR_CMP_COL and pd.isna(v):
                v = None
            c = ws.cell(r, j, v); c.font = Font(name=ARIAL, size=10)
            c.border = Border(left=thin, right=thin, bottom=thin, top=(grp if new else thin))
            if j in numfmt: c.number_format = numfmt[j]
        if new: ws.cell(r,1).font = Font(name=ARIAL, size=10, bold=True)
        prev = row["attribute"]; r += 1
    last = r - 1

    # color scales on the two rate columns (H, I) and lift (J)
    for col in ("H","I","J"):
        ws.conditional_formatting.add(f"{col}{HDR+1}:{col}{last}",
            ColorScaleRule(start_type="min", start_color="F8696B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="max", end_color="63BE7B"))
    for col,w in {"A":37,"B":12,"C":26,"D":9,"E":10,"F":9,"G":12,"H":16,"I":22,"J":14}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"; ws.auto_filter.ref = f"A{HDR}:J{last}"

    # ---- Summary tab: attributes ranked by vertical m0 separation, with the
    #      matching partner-948018 rate at the same top/bottom buckets ----
    sm = wb.create_sheet("Summary")
    recs = []
    for attr, blk in df_res.groupby("attribute"):
        real = blk[~blk["bucket"].isin(["(Missing)","(Other)"])]
        use = real if len(real) >= 2 else blk
        hi = use.loc[use["conversion_rate"].idxmax()]; lo = use.loc[use["conversion_rate"].idxmin()]
        recs.append([attr, blk["type"].iloc[0], blk["bucket"].nunique(),
                     round(use["conversion_rate"].min(),4), round(use["conversion_rate"].max(),4),
                     round(use["conversion_rate"].max()-use["conversion_rate"].min(),4),
                     str(hi["bucket"]), hi[CR_CMP_COL], str(lo["bucket"]), lo[CR_CMP_COL]])
    sh = ["attribute","type","n_buckets","min_conv","max_conv","spread",
          "highest_bucket", f"{cid}_rate_at_high", "lowest_bucket", f"{cid}_rate_at_low"]
    sumdf = pd.DataFrame(recs, columns=sh)
    sumdf = sumdf[sumdf["n_buckets"]>1].sort_values("spread", ascending=False).reset_index(drop=True)

    sm["A1"] = "Attributes ranked by vertical m0 separation (spread = max − min conversion across buckets)"
    sm["A1"].font = Font(name=ARIAL, size=12, bold=True)
    sm["A2"] = (f"Constant attributes omitted. *_rate_at_high/low = partner {cid}'s rate in the vertical's "
                f"best/worst bucket (blank = no {cid} rows there).")
    sm["A2"].font = Font(name=ARIAL, size=9, italic=True, color="555555")
    H2 = 4
    for j,h in enumerate(sh, 1):
        c = sm.cell(H2, j, h); c.font = Font(name=ARIAL, bold=True, color="FFFFFF"); c.fill = navy
        c.alignment = Alignment(horizontal="center")
    ratecols = {4,5,6,8,10}
    for i,row in sumdf.iterrows():
        rr = H2 + 1 + i
        for j,h in enumerate(sh, 1):
            v = row[h]
            if h in (f"{cid}_rate_at_high", f"{cid}_rate_at_low") and pd.isna(v): v = None
            c = sm.cell(rr, j, v); c.font = Font(name=ARIAL, size=10)
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if j in ratecols: c.number_format = "0.0000"
            if j == 3: c.number_format = "0"
    lastS = H2 + len(sumdf)
    if len(sumdf):
        sm.conditional_formatting.add(f"F{H2+1}:F{lastS}",
            ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="63BE7B"))
    for col,w in {"A":37,"B":12,"C":10,"D":10,"E":10,"F":9,"G":24,"H":18,"I":24,"J":18}.items():
        sm.column_dimensions[col].width = w
    sm.freeze_panes = "A5"; sm.auto_filter.ref = f"A{H2}:J{lastS}"

    wb.save(out_path)
    return last - HDR, len(sumdf)


# ======================================================================
# MAIN
# ======================================================================
def main():
    local_csv = os.environ.get("LOCAL_CSV")
    if local_csv:
        df, source = pd.read_csv(local_csv), os.path.basename(local_csv)
    else:
        df = fetch(PARTNER_ID)
        source = "Snowflake query" + (f" (partner_id = {PARTNER_ID})" if PARTNER_ID else " (whole vertical)")

    scope = f"Partner {PARTNER_ID}" if PARTNER_ID is not None else "Whole Vertical"
    res, meta = analyze(df)
    n_detail, n_sum = build_xlsx(res, meta, scope, source, OUT_PATH)
    print(f"{scope}: vertical rows={meta['n_rows']:,} m0={meta['n_m0']:,} overall={meta['overall']:.4f} | "
          f"partner {COMPARE_PARTNER_ID} rows={meta['p_rows']:,} m0={meta['p_m0']:,} "
          f"-> {OUT_PATH} ({n_detail} detail rows, {n_sum} summary attrs)")


if __name__ == "__main__":
    main()